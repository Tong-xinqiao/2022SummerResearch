import torch
import pandas as pd
import numpy as np
import openpyxl
import os


def run(num_neuron, k):
    para = torch.load('model_para_prior' + str(k) + '.pt')
    w_mu = para['hidden.w_mu'].detach().numpy()
    w_rho = para['hidden.w_rho'].detach().numpy()
    w_sigma = np.log(1 + np.exp(w_rho))
    out_w_mu = para['out.w_mu'].detach().numpy()
    out_w_rho = para['out.w_rho'].detach().numpy()

    b_mu = para['hidden.b_mu'].detach().numpy()
    b_rho = para['hidden.b_rho'].detach().numpy()
    b_sigma = np.log(1 + np.exp(b_rho))
    out_b_mu = para['out.b_mu'].detach().numpy()
    out_b_rho = para['out.b_rho'].detach().numpy()

    def write_data():
        # write headers
        for i in range(num_neuron):
            wk_sheet.cell(row=1, column=2 * i + 2, value="Neuron " + str(i + 1))
            wk_sheet.cell(row=1, column=2 * i + 3, value="Neuron " + str(i + 1))
            wk_sheet.cell(row=2, column=2 * i + 2, value="Mean")
            wk_sheet.cell(row=2, column=2 * i + 3, value="Sigma")
        # write mu and sigma in first layer, row by row
        for j in range(len(w_mu[0])):
            wk_sheet.cell(row=j + 3, column=1, value="Factor " + str(j + 1))
            for m in range(num_neuron):
                wk_sheet.cell(row=j + 3, column=2 * m + 2, value=w_mu[m][j])
                wk_sheet.cell(row=j + 3, column=2 * m + 3, value=w_sigma[m][j])
        # write mu and sigma in output layer
        col = 2 * num_neuron + 3  # one column blank
        wk_sheet.cell(row=2, column=col, value='b_mu')
        wk_sheet.cell(row=2, column=col + 1, value='b_sigma')
        wk_sheet.cell(row=2, column=col + 3, value='out_w_wu')
        wk_sheet.cell(row=2, column=col + 4, value='out_w_sigma')
        for i in range(num_neuron):
            wk_sheet.cell(row=3 + i, column=col, value=b_mu[i])
            wk_sheet.cell(row=3 + i, column=col + 1, value=np.log(1 + np.exp(b_rho[i])))
            wk_sheet.cell(row=3 + i, column=col + 3, value=out_w_mu[0][i])
            wk_sheet.cell(row=3 + i, column=col + 4, value=np.log(1 + np.exp(out_w_rho[0][i])))

        wk_sheet.cell(row=2, column=col + 6, value='out_b_mu')
        wk_sheet.cell(row=3, column=col + 6, value=out_b_mu[0])
        wk_sheet.cell(row=2, column=col + 7, value='out_b_sigma')
        wk_sheet.cell(row=3, column=col + 7, value=np.log(1 + np.exp(out_b_rho[0])))

    if os.path.exists("res.xlsx"):
        workbook = openpyxl.load_workbook("res.xlsx")
        wk_sheet = workbook.create_sheet(str(k))
        write_data()
        workbook.save("res.xlsx")
    else:
        workbook = openpyxl.Workbook()
        wk_sheet = workbook.create_sheet(str(k))
        write_data()
        workbook.save("res.xlsx")


for i in range(1):
    run(3, i + 1)
