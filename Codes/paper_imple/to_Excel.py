import torch
import os
import openpyxl


def run(num_neuron, k):
    para = torch.load('model_para' + str(k) + '.pt')
    w = para['fc.weight'].cpu().numpy()
    b = para['fc.bias'].cpu().numpy()
    out_w = para['fc2.weight'].cpu().numpy()
    out_b = para['fc2.bias'].cpu().numpy()

    def write_data():
        # write headers
        for i in range(num_neuron):
            wk_sheet.cell(row=1, column=i + 2, value="Neuron " + str(i + 1))
            wk_sheet.cell(row=2, column=i + 2, value="Weight")
        # write weights in first layer, row by row
        for j in range(len(w[0])):
            wk_sheet.cell(row=j + 3, column=1, value="Factor " + str(j + 1))
            for m in range(num_neuron):
                wk_sheet.cell(row=j + 3, column=m + 2, value=w[m][j])
        # write weight in output layer
        col = num_neuron + 3  # one column blank
        wk_sheet.cell(row=2, column=col, value='bias')
        wk_sheet.cell(row=2, column=col + 2, value='out_w')

        for i in range(num_neuron):
            wk_sheet.cell(row=3 + i, column=col, value=b[i])
            wk_sheet.cell(row=3 + i, column=col + 2, value=out_w[0][i])

        wk_sheet.cell(row=2, column=col + 3, value='out_bias')
        wk_sheet.cell(row=3, column=col + 3, value=out_b[0])

    if os.path.exists("res.xlsx"):
        workbook = openpyxl.load_workbook("res.xlsx")
        wk_sheet = workbook.create_sheet(str(k))
        write_data()
        workbook.save("res.xlsx")
    else:
        workbook = openpyxl.Workbook()
        wk_sheet = workbook.create_sheet(str(k))
        write_data()
        workbook.save("res.xlsx")


for i in range(10):
    run(3, i + 1)
